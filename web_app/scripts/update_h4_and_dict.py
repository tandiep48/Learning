#!/usr/bin/env python3
"""
Production data update: refresh the HSK 4 lessons and the word dictionary.

Two inputs, run together in one transaction:

  1. H4_flag.json  -> lesson_passages + lesson_lines (HSK 4).
     Each H4 passage's lines are fully replaced from the JSON (delete + insert),
     so re-running yields the same result. Passages are upserted with
     hsk_level = 'HSK4'; regular HSK 1-3/5-7 and book passages are never touched.

  2. chinese_dict-*.xlsx -> vocabulary + sematic_diffculty + chinese_stroke_info.
     Upsert by `cn`: existing words are updated in place, new words inserted.
     Nothing is deleted, so passage_vocabulary / user_saved_word foreign keys
     stay intact. sematic_diffculty.word_id is resolved from the resulting
     vocabulary.id (NOT the spreadsheet id); chinese_stroke_info is keyed by cn.
     Duplicate `cn` rows in the sheet collapse to the last occurrence.

Usage (from web_app, using the app's venv):
    python scripts/update_h4_and_dict.py --dry-run
    python scripts/update_h4_and_dict.py --apply
    python scripts/update_h4_and_dict.py --apply --h4 only
    python scripts/update_h4_and_dict.py --apply --dict only
    python scripts/update_h4_and_dict.py --apply --h4-file /path/H4_flag.json --dict-file /path/dict.xlsx

Reads the same DB_* variables as the app from web_app/.env. Point .env at the
target (production) database before running --apply.
"""
import os
import re
import sys
import json
import argparse

from dotenv import load_dotenv
from sqlalchemy import update, insert, bindparam

# Bootstrap: put web_app on the path and load .env BEFORE importing the entity
# layer (entity.database reads the DB_* vars at import time).
_HERE = os.path.dirname(os.path.abspath(__file__))     # web_app/scripts
_WEBAPP = os.path.dirname(_HERE)                        # web_app
load_dotenv(os.path.join(_WEBAPP, ".env"))
sys.path.insert(0, _WEBAPP)

from entity.database import SessionLocal                          # noqa: E402
from entity.passage.entity import LessonPassage                   # noqa: E402
from entity.lesson_line.entity import LessonLine                  # noqa: E402
from entity.vocabulary.entity import Vocabulary                   # noqa: E402
from entity.sematic_difficulty.entity import SemanticDifficulty   # noqa: E402
from entity.chinese_stroke_info.entity import ChineseStrokeInfo   # noqa: E402

# YiChinese/ sits two levels above web_app; the source files live there by default.
_DEFAULT_ROOT = os.path.normpath(os.path.join(_WEBAPP, "..", ".."))
DEFAULT_H4 = os.path.join(_DEFAULT_ROOT, "H4_flag.json")
DEFAULT_DICT = os.path.join(_DEFAULT_ROOT, "chinese_dict-2026-08-16.xlsx")

H4_HSK_LEVEL = "HSK4"
_H_TAG = re.compile(r"^h([1-7])$", re.IGNORECASE)


# --------------------------------------------------------------------------- #
# H4 lessons
# --------------------------------------------------------------------------- #
def load_h4(path):
    """Read H4_flag.json and return its list of passage dicts."""
    # utf-8-sig tolerates a BOM (PowerShell-exported JSON often has one).
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError(f"{path}: expected a list of passages, got {type(data).__name__}")
    return data


def report_h4(session, passages):
    """Return (passage_count, line_count, new_passages, bad_ids)."""
    existing = {
        pid for (pid,) in session.query(LessonPassage.passage_id)
        .filter(LessonPassage.passage_id.like("H4=_%", escape="=")).all()
    }
    line_count = sum(len(p.get("lines", []) or []) for p in passages)
    new_count = sum(1 for p in passages if p["passage_id"] not in existing)
    bad = [p.get("passage_id") for p in passages
           if not str(p.get("passage_id", "")).startswith("H4_")]
    return len(passages), line_count, new_count, bad


def apply_h4(session, passages):
    """Upsert each passage and fully replace its lines."""
    for passage in passages:
        passage_id = passage["passage_id"]

        row = session.get(LessonPassage, passage_id)
        if row is None:
            session.add(LessonPassage(passage_id=passage_id, hsk_level=H4_HSK_LEVEL))
        else:
            row.hsk_level = H4_HSK_LEVEL

        # DELETE runs immediately, so the inserts never clash on the
        # (passage_id, line_id) unique index.
        session.query(LessonLine).filter(LessonLine.passage_id == passage_id).delete(
            synchronize_session=False
        )

        for line in passage.get("lines", []) or []:
            translations = line.get("translations") or {}
            session.add(LessonLine(
                passage_id=passage_id,
                line_id=line.get("line_id"),
                speaker=line.get("speaker"),
                content=line.get("content"),
                pinyin=line.get("pinyin"),
                audio_key=line.get("audio_key"),
                translation_en=translations.get("en"),
                translation_vi=translations.get("vi"),
                tokens=line.get("tokens") or [],
                flag=1 if line.get("flag", 1) else 0,
            ))


# --------------------------------------------------------------------------- #
# Dictionary
# --------------------------------------------------------------------------- #
def _hsk_from_tags(tags):
    """Derive 'HSK<n>' from a comma tag string (e.g. 'h4, verb' -> 'HSK4'), else None."""
    if not tags:
        return None
    for part in str(tags).split(","):
        m = _H_TAG.match(part.strip())
        if m:
            return "HSK" + m.group(1)
    return None


def load_dict(path):
    """Read the dictionary xlsx and return (rows, dropped_dup_cn).

    Rows are deduplicated by `cn` keeping the last occurrence; each row is a
    plain dict of the columns needed by the three target tables.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        idx = {name: i for i, name in enumerate(header)}

        def val(r, name):
            i = idx.get(name)
            return r[i] if i is not None else None

        by_cn = {}
        seen = 0
        for r in it:
            cn = val(r, "cn")
            if cn is None or str(cn).strip() == "":
                continue
            cn = str(cn).strip()
            seen += 1
            by_cn[cn] = {
                "cn": cn,
                "pinyin": val(r, "py"),
                "meaning_en": val(r, "en"),
                "meaning_vn": val(r, "vn"),
                "audio_key": val(r, "audio_key"),
                "hsk_level": _hsk_from_tags(val(r, "tags")),
                "sematic_difficulty": val(r, "sematic_difficulty"),
                "sematic_tags": val(r, "sematic_tags"),
                "zh": val(r, "zh"),
                "total_strokes_cn": val(r, "total_strokes_cn"),
                "total_strokes_zh": val(r, "total_strokes_zh"),
                "strokes_cn": val(r, "strokes_cn"),
                "strokes_zh": val(r, "strokes_zh"),
                "word_length": val(r, "word_length"),
                "strokes_difficult_cn": val(r, "strokes_difficult_cn"),
                "strokes_difficult_cn_norm": val(r, "strokes_difficult_cn_norm"),
                "strokes_difficult_zh": val(r, "strokes_difficult_zh"),
                "strokes_difficult_zh_norm": val(r, "strokes_difficult_zh_norm"),
            }
        return list(by_cn.values()), seen - len(by_cn)
    finally:
        wb.close()


def report_dict(session, rows):
    """Return counts of new vs updated words / stroke rows without writing."""
    xlsx_cns = {r["cn"] for r in rows}
    existing_cns = {cn for (cn,) in session.query(Vocabulary.cn).all()}
    existing_stroke = {cn for (cn,) in session.query(ChineseStrokeInfo.cn).all()}
    vocab_new = len(xlsx_cns - existing_cns)
    stroke_new = len(xlsx_cns - existing_stroke)
    return {
        "words": len(rows),
        "vocab_new": vocab_new,
        "vocab_update": len(rows) - vocab_new,
        "stroke_new": stroke_new,
        "stroke_update": len(rows) - stroke_new,
    }


def apply_dict_rows(session, rows):
    """Upsert vocabulary (by cn), then sematic_diffculty (by word_id) and
    chinese_stroke_info (by cn)."""
    # --- vocabulary: update existing objects in place, add new ones ---
    vocab = {v.cn: v for v in session.query(Vocabulary).all()}
    for r in rows:
        v = vocab.get(r["cn"])
        if v is None:
            v = Vocabulary(cn=r["cn"], source="dictionary")
            session.add(v)
            vocab[r["cn"]] = v
        v.pinyin = r["pinyin"]
        v.meaning_en = r["meaning_en"]
        v.meaning_vn = r["meaning_vn"]
        v.audio_key = r["audio_key"]
        v.source = "dictionary"
        # Only set the HSK level when the sheet gives one; keep any existing value otherwise.
        if r["hsk_level"] is not None:
            v.hsk_level = r["hsk_level"]

    session.flush()  # assign ids to the freshly inserted words
    cn_to_id = {cn: v.id for cn, v in vocab.items()}

    # sematic_diffculty and chinese_stroke_info are upserted with Core statements
    # (executemany). The live sematic_diffculty holds duplicate word_id rows despite
    # the PK in schema.sql, so the ORM's one-row-per-key UPDATE assertion fails; a
    # Core UPDATE just sets every matching row and skips that check.

    sem_tbl = SemanticDifficulty.__table__
    stroke_tbl = ChineseStrokeInfo.__table__

    # --- sematic_diffculty: keyed by vocabulary.id ---
    existing_wid = {wid for (wid,) in session.query(SemanticDifficulty.word_id).all()}
    sem_upd, sem_ins, seen = [], [], set()
    for r in rows:
        wid = cn_to_id.get(r["cn"])
        if wid is None or wid in seen:
            continue
        seen.add(wid)
        payload = {"sematic_difficulty": r["sematic_difficulty"], "tags": r["sematic_tags"]}
        if wid in existing_wid:
            sem_upd.append({"key_wid": wid, **payload})
        else:
            sem_ins.append({"word_id": wid, **payload})
    if sem_upd:
        session.execute(
            update(sem_tbl)
            .where(sem_tbl.c.word_id == bindparam("key_wid"))
            .values(sematic_difficulty=bindparam("sematic_difficulty"), tags=bindparam("tags")),
            sem_upd,
        )
    if sem_ins:
        session.execute(insert(sem_tbl), sem_ins)

    # --- chinese_stroke_info: keyed by cn ---
    stroke_cols = [
        "zh", "total_strokes_cn", "total_strokes_zh", "strokes_cn", "strokes_zh",
        "word_length", "strokes_difficult_cn", "strokes_difficult_cn_norm",
        "strokes_difficult_zh", "strokes_difficult_zh_norm",
    ]
    existing_cn = {cn for (cn,) in session.query(ChineseStrokeInfo.cn).all()}
    str_upd, str_ins, seen = [], [], set()
    for r in rows:
        cn = r["cn"]
        if cn in seen:
            continue
        seen.add(cn)
        payload = {c: r[c] for c in stroke_cols}
        if cn in existing_cn:
            str_upd.append({"key_cn": cn, **payload})
        else:
            str_ins.append({"cn": cn, **payload})
    if str_upd:
        session.execute(
            update(stroke_tbl)
            .where(stroke_tbl.c.cn == bindparam("key_cn"))
            .values({c: bindparam(c) for c in stroke_cols}),
            str_upd,
        )
    if str_ins:
        session.execute(insert(stroke_tbl), str_ins)


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def main():
    parser = argparse.ArgumentParser(
        description="Update HSK 4 lessons and the word dictionary from the source files."
    )
    parser.add_argument("--h4-file", default=DEFAULT_H4, help=f"H4 lesson JSON (default: {DEFAULT_H4})")
    parser.add_argument("--dict-file", default=DEFAULT_DICT, help=f"dictionary xlsx (default: {DEFAULT_DICT})")
    parser.add_argument("--h4", choices=["yes", "only", "no"], default="yes",
                        help="run the H4 lesson refresh (yes/only/no; 'only' skips the dict)")
    parser.add_argument("--dict", choices=["yes", "only", "no"], default="yes",
                        help="run the dictionary update (yes/only/no; 'only' skips H4)")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true", help="report what would change, write nothing")
    group.add_argument("--apply", action="store_true", help="write both updates in one transaction")
    args = parser.parse_args()

    do_h4 = args.h4 != "no" and args.dict != "only"
    do_dict = args.dict != "no" and args.h4 != "only"
    if not do_h4 and not do_dict:
        print("Nothing selected (both sections disabled).")
        sys.exit(1)

    passages = load_h4(args.h4_file) if do_h4 else None
    dict_rows = dropped = None
    if do_dict:
        dict_rows, dropped = load_dict(args.dict_file)

    session = SessionLocal()
    try:
        if do_h4:
            p_count, l_count, p_new, bad = report_h4(session, passages)
            print("H4 lessons:")
            print(f"  file        : {args.h4_file}")
            print(f"  passages    : {p_count}  (new: {p_new}, replace: {p_count - p_new})")
            print(f"  lines       : {l_count}  (fully replaced per passage)")
            if bad:
                print(f"  BAD IDs     : {bad[:5]}{' ...' if len(bad) > 5 else ''}")
            print()

        if do_dict:
            stats = report_dict(session, dict_rows)
            print("Dictionary:")
            print(f"  file        : {args.dict_file}")
            print(f"  words       : {stats['words']}  (dropped {dropped} duplicate-cn rows)")
            print(f"  vocabulary  : new {stats['vocab_new']}, update {stats['vocab_update']}")
            print(f"  stroke_info : new {stats['stroke_new']}, update {stats['stroke_update']}")
            print(f"  sematic     : upserted per word (keyed by vocabulary.id)")
            print()

        if args.dry_run:
            print("dry-run: nothing was written.")
            return

        if do_h4:
            apply_h4(session, passages)
        if do_dict:
            apply_dict_rows(session, dict_rows)

        session.commit()
        print("Applied: updates committed in one transaction.")
    except Exception:
        session.rollback()
        raise
    finally:
        SessionLocal.remove()


if __name__ == "__main__":
    main()
