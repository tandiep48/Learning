#!/usr/bin/env python3
"""
Import grammar rules + context tables from the Grammar_*.xlsx files.

Each workbook has:
  * a 'Grammar' sheet -> grammar_rule
    (GrammarID, Type, Passage_Number, vietnamese, english)
  * one sheet per type-4 table, named by its grammar_id (e.g. 'H1-2-1-1-VN')
    -> grammar_context (grammar_id = sheet name, content_json = list of row dicts).

Re-running replaces all grammar for the HSK levels present in the files (delete by
'H{level}-%' prefix, then insert), so it is idempotent per level. Everything runs in
one transaction.

Usage (from web_app, using the app's venv):
    python scripts/import_grammar.py --dir /path/to/xlsx --dry-run
    python scripts/import_grammar.py --dir /path/to/xlsx --apply

Reads the same DB_* variables as the app from web_app/.env.
"""
import os
import sys
import glob
import json
import argparse

import psycopg2
from psycopg2.extras import Json
from dotenv import load_dotenv
import openpyxl


def connect():
    here = os.path.dirname(os.path.abspath(__file__))          # web_app/scripts
    webapp = os.path.dirname(here)                             # web_app
    load_dotenv(os.path.join(webapp, ".env"))
    return psycopg2.connect(
        host=os.getenv("DB_HOST", "localhost"),
        port=os.getenv("DB_PORT", "5432"),
        dbname=os.getenv("DB_NAME"),
        user=os.getenv("DB_USER"),
        password=os.getenv("DB_PASSWORD"),
    )


def _unique_headers(row):
    """Column headers for a context table. Duplicate header cells get zero-width
    spaces appended so the dict keys stay unique without changing what's displayed."""
    headers = []
    seen = {}
    zwsp = "​"  # zero-width space: keeps duplicate keys unique but visually identical
    for cell in row:
        h = "" if cell is None else str(cell).strip()
        if h in seen:
            seen[h] += 1
            headers.append(h + (zwsp * seen[h]))
        else:
            seen[h] = 0
            headers.append(h)
    return headers


def load_file(path):
    """Return (rules, contexts) parsed from one workbook."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rules = []
    contexts = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if name.strip().lower() == "grammar":
            for r in rows[1:]:  # skip header
                if not r or r[0] is None:
                    continue
                rules.append({
                    "grammar_id": str(r[0]).strip(),
                    "type": int(r[1]) if r[1] is not None else None,
                    "passage_number": int(r[2]) if r[2] is not None else None,
                    "vietnamese_content": r[3],
                    "english_content": r[4],
                })
        else:
            if not rows:
                continue
            headers = _unique_headers(rows[0])
            data = []
            for r in rows[1:]:
                if not r or all(c is None for c in r):
                    continue
                data.append({h: ("" if v is None else str(v)) for h, v in zip(headers, r)})
            contexts[name.strip()] = data
    wb.close()
    return rules, contexts


def main():
    parser = argparse.ArgumentParser(description="Import grammar from Grammar_*.xlsx files.")
    parser.add_argument("--dir", default=".", help="directory holding the Grammar_*.xlsx files (default: .)")
    parser.add_argument("--pattern", default="Grammar_*.xlsx", help="glob for the workbooks (default: Grammar_*.xlsx)")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--dry-run", action="store_true", help="show what would change, write nothing")
    group.add_argument("--apply", action="store_true", help="replace grammar for the found levels in one transaction")
    args = parser.parse_args()

    files = sorted(glob.glob(os.path.join(args.dir, args.pattern)))
    if not files:
        print(f"No files matching '{args.pattern}' in {args.dir}")
        sys.exit(1)

    all_rules = []
    all_contexts = {}
    for path in files:
        rules, contexts = load_file(path)
        all_rules.extend(rules)
        all_contexts.update(contexts)

    # Level prefixes to replace, e.g. 'H1' from grammar_id 'H1-2-1'.
    levels = sorted({r["grammar_id"].split("-")[0] for r in all_rules if r["grammar_id"]})

    print(f"files            : {len(files)}")
    print(f"levels found     : {', '.join(levels)}")
    print(f"grammar_rule rows: {len(all_rules)}")
    print(f"context tables   : {len(all_contexts)}")

    if args.dry_run:
        print("dry-run: nothing was written.")
        return

    conn = connect()
    try:
        with conn.cursor() as cur:
            for level in levels:
                cur.execute("DELETE FROM grammar_rule WHERE grammar_id LIKE %s", (f"{level}-%",))
                cur.execute("DELETE FROM grammar_context WHERE grammar_id LIKE %s", (f"{level}-%",))

            cur.executemany(
                """INSERT INTO grammar_rule
                   (grammar_id, type, passage_number, vietnamese_content, english_content)
                   VALUES (%(grammar_id)s, %(type)s, %(passage_number)s,
                           %(vietnamese_content)s, %(english_content)s)""",
                all_rules,
            )
            cur.executemany(
                "INSERT INTO grammar_context (grammar_id, content_json) VALUES (%s, %s)",
                [(gid, Json(data)) for gid, data in all_contexts.items()],
            )
        conn.commit()
        print(f"Imported {len(all_rules)} grammar rows and {len(all_contexts)} context tables.")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
